import openpyxl
from openpyxl import Workbook

# Define the XPATHs and associated pages and expected outcomes
data = [
    {
        "base_url": "https://example.com",
        "page_type": "Home",
        "search_bar_xp": "//input[@id='js-search-autocomplete']",
        "suggestions_dropdown_xp": "//ul[@id='js-search-items']",
        "suggestion_item_xp": ".//li[contains(@class, 'google-auto-suggestion-list')]",
        "date_picker_xp": "//input[@id='js-date-range-display']",
        "calendar_date_xp": "//td[contains(@class, 'datepicker__month-day--valid')]",
        "continue_button_xp": "//button[@id='js-date-select']",
        "property_title_xp": "",
        "date_availability_xp": "",
        "expected_output": "Search initialized"
    },
    {
        "base_url": "https://example.com",
        "page_type": "Refine",
        "search_bar_xp": "",
        "suggestions_dropdown_xp": "",
        "suggestion_item_xp": "",
        "date_picker_xp": "",
        "calendar_date_xp": "",
        "continue_button_xp": "",
        "property_title_xp": "div.title a",
        "date_availability_xp": "//div[@id='js-date-available'] | //div[@id='js-date-unavailable']",
        "expected_output": "Properties displayed"
    },
    {
        "base_url": "https://example.com",
        "page_type": "Hybrid",
        "search_bar_xp": "",
        "suggestions_dropdown_xp": "",
        "suggestion_item_xp": "",
        "date_picker_xp": "",
        "calendar_date_xp": "",
        "continue_button_xp": "",
        "property_title_xp": "",
        "date_availability_xp": "//div[@id='js-date-available'] | //div[@id='js-date-unavailable']",
        "expected_output": "Dates validated"
    }
]

# Define the output Excel file path
output_file = "workflow_xpaths.xlsx"

# Create a new workbook and select the active sheet
workbook = Workbook()
sheet = workbook.active
sheet.title = "Workflow XPATHs"

# Define the headers
headers = [
    "base_url",
    "page_type",
    "search_bar_xp",
    "suggestions_dropdown_xp",
    "suggestion_item_xp",
    "date_picker_xp",
    "calendar_date_xp",
    "continue_button_xp",
    "property_title_xp",
    "date_availability_xp",
    "expected_output"
]

# Write the headers to the Excel sheet
for col_num, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col_num, value=header)

# Write the data to the Excel sheet
for row_num, entry in enumerate(data, start=2):
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=row_num, column=col_num, value=entry.get(header, ""))

# Save the Excel file
workbook.save(output_file)
print(f"Excel file '{output_file}' has been created successfully.")
